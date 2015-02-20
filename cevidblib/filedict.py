# -*- coding: utf-8 -*-
""" cevidbtool - manage lists on CeviDB groups

Copyright 2014, Nicola Chiapolini v/o Carbon, carbon@cevi.ch

License: GNU General Public License version 3,
         or (at your option) any later version.


filedict.py -- classes to read and write xlsx files

"""
from __future__ import division, print_function, unicode_literals
import openpyxl

def row2person(row):
    """ create data dict from list of Cells

    Parameters
    ----------
    row : list of Cell objects
        cells from a row single row

    Returns
    -------
    data : dict
        dictionary mapping column letter ("A", "CB", ...) to values

    """
    data      = {}
    for c in row:
        data[c.column] = c.value
    return data

def to_coord(col, row):
    """ create coordinate string from column letter/number and row number

    Parameters
    ----------
    col : string or int
        if string column letter ("A", "CB", ...); if integer column number
    row : int
        row number

    Returns
    -------
    coord : string
        string corresponding to coordinate ("A1")

    """
    if type(col) == int:
        col = openpyxl.cell.get_column_letter(col)
    return u"{0}{1}".format(col, row)

class XlsxReader(object):
    """ Class to read xlsx file with person data from CeviDB """

    def __init__(self, config, filename):
        """ initialise reader object

        Parameters
        ----------
        config : cevidblib.config.Settings object
            configuration for file to read
        filename : string
            filename to load (including path)

        """
        self._cfg = config
        self._wb = openpyxl.load_workbook(filename)
        self._start_persons = self._cfg.header_lines+1

        col_key = self._cfg.get_column_key("id")
        self._persons = {}
        row = self._start_persons
        while True:
            pers_id = str(self.active[to_coord(col_key, row)].value)
            try:
                int(pers_id)
            except (ValueError, TypeError):
                break
            selection = "{first}{row}:{last}{row}".format(
               first=self._cfg.column_keys[0],
               last=self._cfg.column_keys[-1], row=row
            )
            person = row2person(self.active.iter_rows(selection).next())
            self._persons[pers_id] = person
            row += 1
        self._start_footer = row

    def cell(self, cell):
        """ read value from cell at given coordinates

        Parameters
        ----------
        cell: string
            cell coordinates
        """
        return self._wb.active[cell].value

    @property
    def active(self):
        """ active sheet """
        return self._wb.active

    @property
    def start_persons(self):
        """ row number of first row containing persons """
        return self._start_persons

    @property
    def start_footer(self):
        """ row number of first row from footer """
        return self._start_footer

    @property
    def persons(self):
        """ dictionary with person data """
        return self._persons


class XlsxWriter(object):
    """ class to write CeviDB data into xlsx File

    while copying header and footer
    """

    def __init__(self, config, new_name, reader):
        """ initialise writer object

        Parameters
        ----------
        config : cevidblib.config.Settings object
            configuration for file to write (and read)
        new_name : string
            filename for new file (including path)
        reader : XlsxReader object
            reader object with  old file loaded
        """
        self._cfg = config
        self._wb = openpyxl.Workbook()
        self._filename = new_name
        self._old_file = reader

        self._start_persons = self._cfg.header_lines+1
        self._start_footer = -1

    def copy_cell(self, src, dst=None):
        """ copy cell content and formating

        Parameters
        ----------
        src : string
            coordinates of source cell
        dst : string
            coordinates of destination cell
        """
        if dst == None:
            dst = src
        self._wb.active[dst].value = self._old_file.active[src].value
        self._wb.active[dst].style = self._old_file.active[src].style

    def copy_header(self):
        """ copy all header cells

        the header cells are defined by the loaded configuration
        """
        for row in range(1,self._start_persons):
            for col in self._cfg.column_keys:
                self.copy_cell(to_coord(col, row))

    def copy_footer(self):
        """ copy all footer cells

        the footer cells are defined by start_footer and
        the number of footer_lines. footer_lines is set in the
        configuration file, start_footer is calculated at the
        end of write_persons() from the number of header an
        person rows.

        Raises
        ------
        RuntimeError
            if start_footer has not been initialised or
            is invalid (is smaller then start_persons)

        """
        if self._start_footer < self._start_persons:
            raise RuntimeError("start of footer not defined")
        old_footer_start = self._old_file.start_footer
        old_footer_rows = range(old_footer_start,
                old_footer_start+self._cfg.footer_lines)
        new_footer_rows = range(self._start_footer,
                self._start_footer+self._cfg.footer_lines)
        for old_row, new_row in zip(old_footer_rows, new_footer_rows):
            for col in self._cfg.column_keys:
                new_cell = to_coord(col, new_row)
                old_cell = to_coord(col, old_row)
                self.copy_cell(old_cell, new_cell)

    def write_persons(self, persons):
        """ write person data to file

        Parameters
        ----------
        persons: dict
            dictionary with person data

        """
        # create list with sorted names and person ids
        # entrys have the form "lastname|id"
        sort_col = self._cfg.get_column_key("last_name")
        sorted_names = []
        for p_key, p_person in persons.items():
            try:
                sorted_names.append(p_person[sort_col]+"|"+p_key)
            except TypeError:
                # entry without a value in the sort_col
                # -> fake a value that gets sorted last
                sorted_names.append("zzz|"+p_key)
        sorted_names.sort()
        # write data for all persons to file (sorted)
        for nr, pers in enumerate(sorted_names):
            row = nr+self._start_persons
            pers_id = pers.split("|")[1]
            pers_data = persons[pers_id]
            # process data columns
            for col in self._cfg.column_keys:
                val = pers_data[col]
                # recreate formulas
                col_config = self._cfg.columns[col]
                if col_config.group == "formula":
                    val = col_config.value.format(row=row, col=col)
                # write data
                cell_coord = to_coord(col, row)
                self._wb.active[cell_coord].value = val
        self._start_footer = self._start_persons+len(persons)

    def fill(self, persons):
        """ helper function to fill a file

        calls the individual functions to copy header and footer
        and write person data

        Parameters
        ----------
        persons: dict
            dictionary with person data

        """
        self.copy_header()
        self.write_persons(persons)
        self.copy_footer()

    def save(self):
        """ save new workbook to file

        additionally freezes view below header and after column set in config
        """
        self._wb.active.freeze_panes = self._wb.active[
                    to_coord(self._cfg.freeze_column, self._start_persons)
                ]
        self._wb.save(self._filename)

